import statistics
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# (backbone, alpha, trial_label) -> {"global": float, "clients": [8 floats]}
data = {
    ("ResNet18", 1.0, "Trial 1 (2026-04-17_06-21-39)"): {
        "global": 0.5166811720336525,
        "clients": [0.5271493212669683, 0.5520581113801453, 0.5450733752620545,
                    0.5260663507109005, 0.5678160919540229, 0.513715710723192,
                    0.5343511450381679, 0.5280172413793104],
    },
    ("ResNet18", 1.0, "Trial 2 (2026-04-17_15-23-51)"): {
        "global": 0.49942062572421786,
        "clients": [0.5, 0.4748201438848921, 0.517814726840855,
                    0.5434782608695652, 0.5444191343963554, 0.5023041474654378,
                    0.5446808510638298, 0.5503355704697986],
    },
    ("ResNet18", 2.0, "Trial 1 (2026-04-17_06-22-03)"): {
        "global": 0.5450359712230216,
        "clients": [0.5671981776765376, 0.5093896713615024, 0.5985576923076923,
                    0.5640495867768595, 0.5763888888888888, 0.5758293838862559,
                    0.5292620865139949, 0.5550755939524838],
    },
    ("ResNet18", 2.0, "Trial 2 (2026-04-17_15-20-54)"): {
        "global": 0.5249351398097435,
        "clients": [0.5673289183222958, 0.5379146919431279, 0.49563318777292575,
                    0.5439024390243903, 0.5460992907801419, 0.5789473684210527,
                    0.5216346153846154, 0.5111111111111111],
    },
    ("ResNet50", 0.1, "Trial 1 (2026-04-17_06-36-31)"): {
        "global": 0.6317930834059866,
        "clients": [0.7137014314928425, 0.5796766743648961, 0.6582633053221288,
                    0.6569506726457399, 0.6529284164859002, 0.645320197044335,
                    0.6693989071038251, 0.6853002070393375],
    },
    ("ResNet50", 0.1, "Trial 2 (2026-04-17_19-02-14)"): {
        "global": 0.6758801280186209,
        "clients": [0.7480519480519481, 0.6642857142857143, 0.6802884615384616,
                    0.6869918699186992, 0.7282608695652174, 0.7252475247524752,
                    0.6857142857142857, 0.6962962962962963],
    },
}

# group by (backbone, alpha)
groups = {}
for (bb, alpha, trial), v in data.items():
    groups.setdefault((bb, alpha), []).append((trial, v))

def pct(x):
    return x * 100.0

def fmt(mean, std):
    return f"{mean:.2f} +- {std:.2f}"

wb = Workbook()

# ---------- Sheet 1: Summary ----------
ws = wb.active
ws.title = "Summary"
headers = ["Backbone", "Alpha", "# Trials",
           "Global Model Acc (%) [mean +- std over trials]",
           "Personalized Model Acc (%) [mean +- std over clients x trials]"]
ws.append(headers)
for c in ws[1]:
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor="DDEBF7")

for (bb, alpha), trials in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1])):
    global_vals = [pct(v["global"]) for _, v in trials]
    client_vals = [pct(c) for _, v in trials for c in v["clients"]]
    g_mean = statistics.mean(global_vals)
    g_std = statistics.stdev(global_vals) if len(global_vals) > 1 else 0.0
    p_mean = statistics.mean(client_vals)
    p_std = statistics.stdev(client_vals) if len(client_vals) > 1 else 0.0
    ws.append([bb, alpha, len(trials), fmt(g_mean, g_std), fmt(p_mean, p_std)])

for col, width in zip("ABCDE", [12, 8, 10, 45, 55]):
    ws.column_dimensions[col].width = width
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(horizontal="center", vertical="center")

# ---------- Sheet 2: Per-Trial Detail ----------
ws2 = wb.create_sheet("Per-Trial Detail")
hdr2 = ["Backbone", "Alpha", "Trial",
        "Global Model Acc (%)",
        "Personalized Acc (%) [mean +- std over 8 clients]"]
ws2.append(hdr2)
for c in ws2[1]:
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor="DDEBF7")

for (bb, alpha), trials in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1])):
    for trial_label, v in trials:
        g = pct(v["global"])
        client_pct = [pct(c) for c in v["clients"]]
        c_mean = statistics.mean(client_pct)
        c_std = statistics.stdev(client_pct)
        ws2.append([bb, alpha, trial_label, f"{g:.2f}", fmt(c_mean, c_std)])

for col, width in zip("ABCDE", [12, 8, 36, 22, 45]):
    ws2.column_dimensions[col].width = width
for row in ws2.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(horizontal="center", vertical="center")

# ---------- Sheet 3: Raw per-client ----------
ws3 = wb.create_sheet("Raw Per-Client")
hdr3 = ["Backbone", "Alpha", "Trial"] + [f"Client {i} (%)" for i in range(8)] + ["Global (%)"]
ws3.append(hdr3)
for c in ws3[1]:
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor="DDEBF7")

for (bb, alpha), trials in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1])):
    for trial_label, v in trials:
        row = [bb, alpha, trial_label] + [f"{pct(c):.2f}" for c in v["clients"]] + [f"{pct(v['global']):.2f}"]
        ws3.append(row)

ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 8
ws3.column_dimensions["C"].width = 36
for col in "DEFGHIJKL":
    ws3.column_dimensions[col].width = 13
for row in ws3.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(horizontal="center", vertical="center")

out = "/root/FedTesBABU/Stanford_dog_results.xlsx"
wb.save(out)
print("Saved:", out)

# Also print summary
print("\n=== Summary ===")
for (bb, alpha), trials in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1])):
    global_vals = [pct(v["global"]) for _, v in trials]
    client_vals = [pct(c) for _, v in trials for c in v["clients"]]
    g_mean = statistics.mean(global_vals); g_std = statistics.stdev(global_vals)
    p_mean = statistics.mean(client_vals); p_std = statistics.stdev(client_vals)
    print(f"{bb:9s} alpha={alpha}  Global: {fmt(g_mean,g_std)}  Personalized: {fmt(p_mean,p_std)}  (n_trials={len(trials)})")
